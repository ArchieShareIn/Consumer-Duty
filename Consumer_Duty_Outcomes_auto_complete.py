import pandas as pd
from datetime import datetime, timedelta
import sqlite3
import os
import sys

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, numbers
from io import BytesIO

import streamlit as st

def last_day_of_month(date):
    next_month = date.replace(day=28) + timedelta(days=4)  # Start from the 28th of current month
    return next_month - timedelta(days=next_month.day)

st.title("Consumer Duty Auto Complete")
years = list(range(2000, 2081))
months = list(range(1, 13))

default_start_month_index = months.index(7) # If a new excel document is used please change the start date here.
default_start_year_index = years.index(2023)

selected_start_month = st.selectbox("Select Month at the start of excel file", months, index=default_start_month_index)
selected_start_year = st.selectbox("Select Year at the start of excel file", years, index=default_start_year_index)

today = datetime.now().date()
first = today.replace(day=1)
monthToChange = first - timedelta(days=1)

selected_month = st.selectbox("Select Month to analyse", months, index=months.index(monthToChange.month))
selected_year = st.selectbox("Select Year to analyse", years, index=years.index(monthToChange.year))

monthToChange = monthToChange.replace(month=selected_month, year=selected_year)
monthToChange = last_day_of_month(monthToChange)

# dateFormat = r'\d{2}/\d{2}/\d{4}'
# startDateExcelFile = datetime.strptime("01/12/2023", '%d/%m/%Y')
if selected_start_month < 10:
    startDateExcelFile = datetime.strptime(f"01/0{selected_start_month}/{selected_start_year}", '%d/%m/%Y')
else:
    startDateExcelFile = datetime.strptime(f"01/{selected_start_month}/{selected_start_year}", '%d/%m/%Y')

# csvLocation = st.text_input("Enter csv files address for AR registed users (remember the .csv)")
# excelFileLocation = st.text_input("Enter Excel files address for the AR calender (remember the .xlsx)")

csvFile = st.file_uploader("Choose the AR registered user CSV file", type="csv")
excelFile = st.file_uploader("Choose the Excel Consumer Duty file", type="xlsx")

sheet_name = st.text_input("Enter the name of the sheet that should be changed within the Excel File")

submitted = st.button("Submit")

def do_analysis():
    #Read the csv file into a pandas dataframe
    df = pd.read_csv(csvFile, skiprows=1, encoding='ISO-8859-1', index_col=False)
    df.columns = df.columns.str.replace(" ", "_")
    df.replace('-', '', inplace=True)

    slectedColumns = ['Investor_Id', 'Create_Date', 'Date_Last_Took_App_Test', 'Number_App_Test_Fails', 'Number_App_Test_Passes', 'Last_Login', 'Last_Investment', 
        'Categorisation', 'Email_Address', 'Email_Confirmed', 'Nationality', 'Resident_Country', 'Test_Investor', 'User_Kyc_Status', 'Pep', 'Vulnerable_Customer']
    monitoring_df=df[slectedColumns].copy()

    # print(monitoring_df)

    #monitoring_df.head()

    #View first 5 rows
    # print(monitoring_df.dtypes)
    monitoring_df['Investor_Id'] = monitoring_df['Investor_Id'].astype(int)
    monitoring_df['Create_Date'] = pd.to_datetime(monitoring_df['Create_Date'],format="%d/%m/%Y")
    monitoring_df['Date_Last_Took_App_Test'] = pd.to_datetime(monitoring_df['Date_Last_Took_App_Test'], format="%d/%m/%Y")
    monitoring_df['Last_Login'] = pd.to_datetime(monitoring_df['Last_Login'], format="%d/%m/%Y")
    monitoring_df['Last_Investment'] = pd.to_datetime(monitoring_df['Last_Investment'], format="%d/%m/%Y")
    monitoring_df['Number_App_Test_Fails'] = monitoring_df['Number_App_Test_Fails'].astype(int)
    monitoring_df['Number_App_Test_Passes'] = monitoring_df['Number_App_Test_Passes'].astype(int)
    # print(monitoring_df.dtypes)
    database = 'databaseConsumer.db'

    if os.path.exists(database):
        os.remove(database)
    # Creating SQLite database
    conn = sqlite3.connect(database)
    cursor = conn.cursor()

    # # Create SQLite table - Table Definition
    create_table = '''CREATE TABLE IF NOT EXISTS ARMonitoringData(
                    Investor_Id INTEGER PRIMARY KEY NOT NULL,
                    Create_Date datetime NOT NULL,
                    Date_Last_Took_App_Test datetime,
                    Last_Login datetime,
                    Last_Investment datetime,
                    Number_App_Test_Fails INTEGER,
                    Number_App_Test_Passes INTEGER,
                    Categorisation VARCHAR(40),
                    Email_Address VARCHAR(320),
                    Email_Confirmed VARCHAR(5),
                    Nationality VARCHAR(40),
                    Resident_Country VARCHAR(40),
                    Test_Investor VARCHAR(5),
                    User_Kyc_Status VARCHAR(40),
                    Pep VARCHAR(5),
                    Vulnerable_Customer VARCHAR(5));
                    '''

    # Creating the table into our database
    cursor.execute(create_table)

    # insert the data from the pandas DataFrame into the SQLite table
    monitoring_df.to_sql('ARMonitoringData', conn, if_exists='replace', index = False)

    # print(pd.read_sql("""SELECT COUNT(*) 
    #     FROM ARMonitoringData 
    #     WHERE Email_Address LIKE '%test%' OR Email_Address LIKE '%%sharein%';""", conn).iloc[0, 0])

    def removeTesters():
        deletion_query = """DELETE FROM ARMonitoringData 
        WHERE Email_Address LIKE '%test%' OR Email_Address LIKE '%%sharein%';"""

        conn.execute(deletion_query)

    removeTesters()

    workbook = openpyxl.load_workbook(excelFile)
    sheet = workbook[sheet_name]
    # sheet = workbook.active
    border = Border(left=Side(style='thin', color='000000'), 
                    right=Side(style='thin', color='000000'), 
                    top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'))

    if (monthToChange.year*12 + monthToChange.month >= startDateExcelFile.year*12 + startDateExcelFile.month):
        COLUMN_OFFSET = 2
        ASCII_OF_A = 65

        letterForMonthInt = (monthToChange.year - startDateExcelFile.year)*12 + (monthToChange.month - startDateExcelFile.month)
        letterForMonthInt += ASCII_OF_A + COLUMN_OFFSET
        if letterForMonthInt < ASCII_OF_A + 26: # If the column is one character long.
            cellColumnForMonth = chr(letterForMonthInt)
        elif letterForMonthInt < ASCII_OF_A + 26*26: # If the column is two characters long.
            cellColumnForMonth = str(chr(int((letterForMonthInt-ASCII_OF_A)/26) + ASCII_OF_A - 1)) + str(chr(int((letterForMonthInt-ASCII_OF_A)%26) + ASCII_OF_A))
        else:
            print("Error: The month entered is two far into the future. Date entered is more than 56 years ahead of the start date")
    else:
        print(f"Error: The inputted date is before {startDateExcelFile}.")
        sys.exit(1)

    def paintCell(cell, colour):
        if colour == "green":
            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        elif colour == "yellow":
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif colour == "amber":
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        elif colour == "red":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif colour == "dark red":
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        elif colour == "white":
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        elif colour == "grey":
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        else:
            print("Error: Invalid colour. The colour must be one of green, yellow, amber, red, dark red, white, and grey.")
            sys.exit(1)
        cell.border = border

    def writeToCellPercent(cell, value):
        cell.value = value * 100
        cell.number_format = "0.0"

    def writeToCellNum(cell, value):
        cell.value = value
        cell.number_format = numbers.FORMAT_NUMBER

    def checkIfZero(num):
        if num == 0:
            print("Error: You are trying to divide by zero when calculating one of the percentages. This is because no people have registered within a time frame. " +
            "The month entered has not happened yet or the csv is empty (probaly one of these reasons).")
            sys.exit(1)

    def convert_to_yyyy_mm(date):
        month = date.month
        strMonth = str(month)
        if month <= 9:
            strMonth = "0" + strMonth
        
        return str(date.year) + "-" + strMonth

    # PSO02

    # print(df[(df["Email_Confirmed"] != "Yes") & (df["Email_Confirmed"] != "No")]["Email_Confirmed"])


    num_not_passed_app_test_with_email = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Create_Date) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND Email_Confirmed = 'Yes' AND Number_App_Test_Passes = 0;", 
        conn).iloc[0, 0]

    num_registered_with_email = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Create_Date) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND Email_Confirmed = 'Yes';", conn).iloc[0, 0]

    checkIfZero(num_registered_with_email)

    percent_not_passed_PSO02 = num_not_passed_app_test_with_email / num_registered_with_email

    cellPos = cellColumnForMonth + "2"
    cell = sheet[cellPos]
    paintCell(cell, "grey")
    writeToCellNum(cell, num_registered_with_email)

    if (percent_not_passed_PSO02 < 0):
        print("Error: The percentage of investors in the last month that have not passed the app test cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "3"
        cell = sheet[cellPos]
        if (percent_not_passed_PSO02 < 0.2):
            paintCell(cell, "red")
        else:
            paintCell(cell, "white")
        writeToCellPercent(cell, percent_not_passed_PSO02)


    # PSO03
    # percentage Of users who created an account and attempted the app test in the monitoring month, percentage who passed the app test on the first attempt.
    # Aim: to understand the proportion of new investors on the platform who appear to have a ‘confident’ grasp of the risks and products.

    num_passed_first_try = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Create_Date) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND strftime('%Y-%m', Date_Last_Took_App_Test) = '""" + 
        convert_to_yyyy_mm(monthToChange) + "' AND Number_App_Test_Passes != 0 AND Number_App_Test_Fails = 0;", conn).iloc[0, 0]

    num_attempted = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Create_Date) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND strftime('%Y-%m', Date_Last_Took_App_Test) = '""" + 
        convert_to_yyyy_mm(monthToChange) + "' AND (Number_App_Test_Passes != 0 OR Number_App_Test_Fails != 0);", conn).iloc[0, 0]

    checkIfZero(num_attempted)

    percent_passed_first_try_PSO03 = num_passed_first_try / num_attempted

    cellPos = cellColumnForMonth + "4"
    cell = sheet[cellPos]
    paintCell(cell, "grey")
    writeToCellNum(cell, num_attempted)

    if (percent_passed_first_try_PSO03 < 0):
        print("Error: The percentage of investors in the last month that have attempted and passed first try the app test cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "5"
        cell = sheet[cellPos]
        if (percent_passed_first_try_PSO03 < 0.55 or percent_passed_first_try_PSO03 > 0.95):
            paintCell(cell, "red")
        else:
            paintCell(cell, "white")
        writeToCellPercent(cell, percent_passed_first_try_PSO03)

    # PSO04
    # Of users who created an account and attempted the app test in the monitoring month, percentage who passed after failing at least once.
    # Aim: to identify potential risk to ShareIn of baseline vs possible increase in the investor base who may be ‘borderline’ 
    # in their understanding of the products and risks and who may therefore need more support or be more likely to complain in future. 

    num_passed_not_first_try = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Create_Date) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND strftime('%Y-%m', Date_Last_Took_App_Test) = '""" + 
        convert_to_yyyy_mm(monthToChange) + "' AND Number_App_Test_Passes != 0 AND Number_App_Test_Fails != 0;", conn).iloc[0, 0]

    percent_passed_not_first_try_PSO04 = num_passed_not_first_try / num_attempted

    if (percent_passed_not_first_try_PSO04 < 0):
        print("Error: The percentage of investors in the last month that have attempted and passed not first try the app test cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "6"
        cell = sheet[cellPos]
        if (percent_passed_not_first_try_PSO04 > 0.15):
            paintCell(cell, "red")
        else:
            paintCell(cell, "white")
        writeToCellPercent(cell, percent_passed_not_first_try_PSO04)

    # CUO06

    # Disengaged investor: Average number of days since last platform login, where customer made an investment in the past 6 months.
    # Aim: identify customers who are disengaged and potentially at higher risk of harm associated with lack of information.



    # print(df["Last_Login"])
    # """valid_rows = monitoring_df.dropna(subset=['Last_Investment', 'Last_Login'])

    # if monthToChange.month > 6:
    #     average_days_last_login = ((monthToChange - valid_rows[(valid_rows["Last_Investment"].dt.month >= monthToChange.month - 6) &
    #     (valid_rows["Last_Investment"].dt.month <= monthToChange.month) & 
    #     (valid_rows["Last_Investment"].dt.year == monthToChange.year)]["Last_Login"]).dt.days).mean()
    # else:
    #     average_days_last_login = ((monthToChange - valid_rows[(valid_rows["Last_Investment"].dt.month >= 12 - monthToChange.month) &
    #     (valid_rows["Last_Investment"].dt.month <= monthToChange.month) & 
    #     (valid_rows["Last_Investment"].dt.year == monthToChange.year - 1)]["Last_Login"]).dt.days).mean()

    # print(average_days_last_login)"""

    last_6_month_invest = pd.read_sql( 
        f"""SELECT *
        FROM ARMonitoringData
        WHERE (strftime('%Y', Last_Investment) = '{monthToChange.year}' AND strftime('%m', Last_Investment) >= '{monthToChange.month - 6}' 
        AND strftime('%m', Last_Investment) <= '{monthToChange.month}') OR (strftime('%Y', Last_Investment) = '{monthToChange.year - 1}' AND 
        strftime('%m', Last_Investment) >= '{12 - monthToChange.month}' AND strftime('%m', Last_Investment) <= '{monthToChange.month}');""", conn)

    # average_days_last_login = ((monthToChange - pd.to_datetime(last_6_month_invest["Last_Login"],format="mixed", dayfirst=True)).dt.days).mean()
    timeStampDate = pd.Timestamp(monthToChange)
    last_6_month_invest["Parsed_Last_Login"] = pd.to_datetime(last_6_month_invest["Last_Login"], format='mixed', dayfirst=True)
    days_difference = (timeStampDate - last_6_month_invest["Parsed_Last_Login"]).dt.days
    positive_days_difference = days_difference[days_difference >= 0]
    average_days_last_login = positive_days_difference.mean()

    if (average_days_last_login < 0):
        print("Error: The average days since last login cannot be negative. The csv of AR data contains dates ahead of the month to change.")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "11"
        cell = sheet[cellPos]
        if (average_days_last_login > 120):
            paintCell(cell, "red")
        else:
            paintCell(cell, "white")
        writeToCellNum(cell, average_days_last_login)
        
    conn.commit()
    conn.close()
    edited_excel_file = BytesIO()
    workbook.save(edited_excel_file)
    edited_excel_file.seek(0)

    st.download_button(
            label="Download Edited Excel File",
            data=edited_excel_file,
            file_name='Consumer-Duty.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

if submitted:
    do_analysis()